#!/usr/bin/env python3
"""Reviewer point 2: add an explicit methylation-compartment flag to Supplementary Table S2 and
run the promoter-only sensitivity analysis.

Each of the 82 inverse-concordant genes is classified by the source of its methylation call:
  - mCSEA promoter-region enrichment (promoter-anchored, consistent with the silencing model)
  - the gene-level Stouffer composite over promoter AND gene-body probes (compartment-ambiguous)
Genes supported only by the composite cannot be attributed to promoter methylation and are
therefore flagged; the sensitivity analysis reports how many survive a promoter-only definition.
"""
import pandas as pd, numpy as np, shutil, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

R="__MS_GEO_ROOT__/Methylation/results"
SRC="__MS_GEO_ROOT__/Poster_v2/Supplementary_Tables_IJMS.xlsx"
DST="__MS_GEO_ROOT__/Poster_v2/Supplementary_Tables_IJMS_v2.xlsx"

src=pd.read_csv(f"{R}/INVERSE_CONCORDANT_by_gene_by_meth_source.tsv",sep="\t")
src.columns=["gene","stouffer","mcsea"]
prom=pd.read_csv(f"{R}/06_mCSEA_promoter.tsv",sep="\t")
body=pd.read_csv(f"{R}/06_mCSEA_gene_body.tsv",sep="\t")
pp=prom.groupby("gene").padj.min(); bp=body.groupby("gene").padj.min()

def classify(r):
    has_p=pd.notna(r.mcsea); has_c=pd.notna(r.stouffer)
    if has_p and has_c: return "Promoter-confirmed (mCSEA + composite)"
    if has_p:           return "Promoter-only (mCSEA)"
    return "Composite only (promoter + gene body)"
src["compartment"]=src.apply(classify,axis=1)
src["survives_promoter_only"]=np.where(src.mcsea.notna(),"Yes","No")
src["mcsea_strata"]=src.mcsea.fillna("-")
src["mcsea_padj"]=src.gene.map(pp)
src["genebody_padj"]=src.gene.map(bp)

T1=["ITGB2","CD79B","IKZF1"]
AUX=["CASP6","CASP8","DGKQ","MX1","IFIT1","NUP210","RUNX3","LXN","SH3BP4"]
T2N=["CTSZ","CHL1","ICAM1","FOXP3","ITGAL"]

print("="*78); print("REVIEWER POINT 2 - PROMOTER-ONLY SENSITIVITY ANALYSIS"); print("="*78)
n=len(src)
print(f"inverse-concordant discovery pool           : {n} genes")
for lab,q in [("promoter-confirmed (mCSEA + composite)",src.compartment.str.startswith("Promoter-confirmed")),
              ("promoter-only (mCSEA, no composite)   ",src.compartment.str.startswith("Promoter-only")),
              ("composite only (promoter + gene body) ",src.compartment.str.startswith("Composite"))]:
    print(f"  {lab}: {q.sum():3d}  ({q.sum()/n*100:.0f}%)")
surv=src[src.survives_promoter_only=="Yes"]
print(f"\nSURVIVE a promoter-only definition: {len(surv)}/{n} ({len(surv)/n*100:.0f}%)")
print(f"  genes: {sorted(surv.gene)}")
for lab,lst in [(f"Tier-1 ({len(T1)})",T1),(f"Tier-2 auxiliary ({len(AUX)})",AUX),("Tier-2 non-concordant (5)",T2N)]:
    s=src[src.gene.isin(lst)]
    ok=sorted(s[s.survives_promoter_only=="Yes"].gene); no=sorted(s[s.survives_promoter_only=="No"].gene)
    print(f"\n{lab}: {len(ok)}/{len(s)} survive")
    print(f"   survive : {ok}")
    print(f"   dropped : {no}")
print("\nmCSEA promoter padj for the named candidates:")
for g in T1+["HLA-E"]+AUX:
    s=src[src.gene==g]
    if not len(s): print(f"   {g:8s} not in the 82-gene pool"); continue
    r=s.iloc[0]
    pv=f"{r.mcsea_padj:.4f}" if pd.notna(r.mcsea_padj) else "not in mCSEA table"
    print(f"   {g:8s} {r.compartment:38s} promoter padj={pv}")

# ---------------- write updated S2 ----------------
shutil.copy(SRC,DST)
wb=load_workbook(DST); ws=wb["Supplementary_Table_S2"]
hdr_row=3                                   # header is on row 3 (title row 1, blank row 2)
# Keep reviewer-driven candidate reclassifications persistent if this generator is rerun.
for i in range(hdr_row+1,ws.max_row+1):
    if ws.cell(row=i,column=1).value in {"LXN", "SH3BP4"}:
        ws.cell(row=i,column=2,value="2 — Tier-2 auxiliary inverse-concordant")
first=ws.max_column+1
newcols=[("Methylation compartment","compartment"),
         ("mCSEA promoter strata","mcsea_strata"),
         ("mCSEA promoter padj (min)","mcsea_padj"),
         ("mCSEA gene-body padj (min)","genebody_padj"),
         ("Survives promoter-only filter","survives_promoter_only")]
idx=src.set_index("gene")
HDR=PatternFill("solid",start_color="1F4E78"); HF=Font(name="Arial",bold=True,color="FFFFFF",size=10)
YES=PatternFill("solid",start_color="C6EFCE"); NO=PatternFill("solid",start_color="FCE4E4")
for k,(title,col) in enumerate(newcols):
    j=first+k; ws.cell(row=hdr_row,column=j,value=title)
    c=ws.cell(row=hdr_row,column=j); c.fill=HDR; c.font=HF
    c.alignment=Alignment(horizontal="center",wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width=26
    for i in range(hdr_row+1,ws.max_row+1):
        g=ws.cell(row=i,column=1).value
        if g is None or g not in idx.index: continue
        v=idx.loc[g,col]
        cell=ws.cell(row=i,column=j,value=(None if pd.isna(v) else v))
        cell.font=Font(name="Arial",size=10)
        if isinstance(v,float): cell.number_format="0.0000"
        if col=="survives_promoter_only": cell.fill=YES if v=="Yes" else NO
ws.cell(row=1,column=1,value=(str(ws.cell(row=1,column=1).value).rstrip()+
    " Methylation-compartment columns indicate whether each gene's inverse-concordant methylation call"
    " was supported by mCSEA promoter-region enrichment (promoter-anchored) or only by the gene-level"
    " Stouffer composite over promoter and gene-body probes."))
wb.save(DST)
print(f"\nwrote {DST}  (+{len(newcols)} columns on Supplementary_Table_S2)")
