#!/usr/bin/env python3
r"""Rebuild Supplementary Table S2 for the current 94-gene inverse-concordant pool.

WHY THIS EXISTS. S2 was last written for the 82-gene pool, before the expression-matrix repair
raised the pool to 94, and its tier column still placed CD79B in Tier-1. Two scripts had since
bolted columns onto that stale sheet (update_S2_promoter_flag.py, update_S2_weighting.py), so the
compartment flag a reviewer asked for existed only for 82 of the 94 genes. This script rebuilds the
sheet from the result tables in one pass instead, so the gene list, the tiers and every derived
column come from the same current inputs.

WHAT THE COMPARTMENT COLUMNS ANSWER. The gene-level methylation statistic aggregates promoter and
gene-body probes, and gene-body methylation often correlates POSITIVELY with transcription - the
opposite of the promoter-silencing model the inverse-concordant filter assumes. A gene can
therefore pass the filter on gene-body signal alone, which would be spurious under that model. Each
gene is classified by what supports its methylation call:

  Promoter-confirmed (mCSEA + composite)   promoter-region enrichment AND the gene-level composite
  Promoter-only (mCSEA)                    promoter-region enrichment, no composite call
  Composite only (promoter + gene body)    compartment-ambiguous; cannot be attributed to promoter

and "Survives promoter-only filter" states whether the gene would remain if the filter were
restricted to mCSEA-significant promoter regions.

WEIGHTING COLUMNS are taken unchanged from the 05_combined (pan-tissue) stratum of
15_genelevel_weighting_corrected.tsv, which is the stratum the manuscript reports. The reported
effect size stays the unweighted probe mean; the 1/SE-weighted and inverse-variance estimators and
the probe range sit beside it so the sensitivity to the weighting choice is visible (Reviewer 1
point 8). Genes with no probe in the combined stratum are left blank rather than imputed.
"""
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = "__MS_GEO_ROOT__"
R = f"{ROOT}/Methylation/results"
SRC = f"{ROOT}/Poster_v2/Supplementary_Tables_IJMS_v2.xlsx"
DST = f"{ROOT}/Poster_v2/Supplementary_Tables_IJMS_v3.xlsx"

TIER1 = ["ITGB2", "IKZF1"]
AUX = ["CD79B", "LXN", "HLA-E", "CASP6", "CASP8", "DGKQ",
       "MX1", "IFIT1", "NUP210", "RUNX3", "SH3BP4"]

# ---------------------------------------------------------------- inputs
g = pd.read_csv(f"{R}/INVERSE_CONCORDANT_by_gene.tsv", sep="\t")
src = pd.read_csv(f"{R}/INVERSE_CONCORDANT_by_gene_by_meth_source.tsv", sep="\t")
src.columns = ["gene", "stouffer", "mcsea"]
prom = pd.read_csv(f"{R}/06_mCSEA_promoter.tsv", sep="\t").groupby("gene").padj.min()
body = pd.read_csv(f"{R}/06_mCSEA_gene_body.tsv", sep="\t").groupby("gene").padj.min()

w = pd.read_csv(f"{R}/15_genelevel_weighting_corrected.tsv", sep="\t")
w = w[w.stratum == "05_combined"].set_index("gene")
w["range_txt"] = [f"{lo:+.3f} … {hi:+.3f}" if n > 1 else "single probe"
                  for lo, hi, n in zip(w.min_logFC, w.max_logFC, w.n_probes)]

d = g.merge(src, on="gene", how="left")
assert len(d) == len(g) == 94, f"pool {len(g)}, merged {len(d)}"


def tier(x):
    if x in TIER1:
        return "1 — Inverse-concordant Tier-1"
    if x in AUX:
        return "2 — Tier-2 auxiliary (inverse-concordant)"
    return "Extended inverse-concordant set"


def compartment(r):
    has_p, has_c = pd.notna(r.mcsea), pd.notna(r.stouffer)
    if has_p and has_c:
        return "Promoter-confirmed (mCSEA + composite)"
    if has_p:
        return "Promoter-only (mCSEA)"
    return "Composite only (promoter + gene body)"


d["Tier"] = d.gene.map(tier)
d["compartment"] = d.apply(compartment, axis=1)
d["survives"] = np.where(d.mcsea.notna(), "Yes", "No")
d["prom_padj"] = d.gene.map(prom)
d["body_padj"] = d.gene.map(body)
for c in ("n_probes", "mean_logFC", "se_logFC", "ivw_logFC", "range_txt", "FDR_A"):
    d[c] = d.gene.map(w[c]) if c in w else np.nan

order = {"1 — Inverse-concordant Tier-1": 0,
         "2 — Tier-2 auxiliary (inverse-concordant)": 1,
         "Extended inverse-concordant set": 2}
d = d.sort_values(["Tier", "gene"], key=lambda s: s.map(order) if s.name == "Tier" else s)

# ---------------------------------------------------------------- sensitivity analysis
print("=" * 78)
print("PROMOTER-ONLY SENSITIVITY ANALYSIS — 94-gene inverse-concordant pool")
print("=" * 78)
for lab, m in [("promoter-confirmed (mCSEA + composite)", d.compartment.str.startswith("Promoter-confirmed")),
               ("promoter-only (mCSEA, no composite)   ", d.compartment.str.startswith("Promoter-only")),
               ("composite only (promoter + gene body) ", d.compartment.str.startswith("Composite"))]:
    print(f"  {lab}: {m.sum():>3} / 94  ({100*m.mean():.0f}%)")
surv = d[d.survives == "Yes"]
print(f"\n  survive a promoter-only filter        : {len(surv)} / 94")
print(f"  genes: {', '.join(sorted(surv.gene))}")
for lab, lst in [("Tier-1", TIER1), ("Tier-2 auxiliary", AUX)]:
    s = d[d.gene.isin(lst)]
    y = sorted(s.loc[s.survives == "Yes", "gene"])
    print(f"\n  {lab} ({len(lst)}): {len(y)} survive promoter-only -> {', '.join(y) or 'none'}")
    print(f"     composite-only: {', '.join(sorted(s.loc[s.survives=='No','gene']))}")

# ---------------------------------------------------------------- write sheet
COLS = [("Gene", "gene", None, 12), ("Tier", "Tier", None, 34),
        ("N cohort pairings", "n_pairings", "0", 12),
        ("Best RNA FDR", "best_rna_fdr", "0.00E+00", 13),
        ("Best RNA log2FC (MS vs ctrl)", "best_rna_fc", "0.0000", 15),
        ("Best methylation FDR", "best_meth_fdr", "0.00E+00", 13),
        ("Best methylation log2FC (MS vs ctrl)", "best_meth_fc", "0.0000", 15),
        ("Direction (RNA / methylation)", "direction", None, 20),
        ("Evidence sources", "sources", None, 46),
        ("Cross-omics (RNA+meth same gene)", "is_cross_omics", None, 14),
        ("Methylation compartment", "compartment", None, 30),
        ("mCSEA promoter strata", "mcsea", None, 24),
        ("mCSEA promoter padj (min)", "prom_padj", "0.0000", 14),
        ("mCSEA gene-body padj (min)", "body_padj", "0.0000", 14),
        ("Survives promoter-only filter", "survives", None, 14),
        ("CpG probes (n)", "n_probes", "0", 11),
        ("Methylation logFC (unweighted mean, reported)", "mean_logFC", "0.0000", 18),
        ("Methylation logFC (1/SE-weighted)", "se_logFC", "0.0000", 16),
        ("Methylation logFC (inverse-variance)", "ivw_logFC", "0.0000", 16),
        ("Probe logFC range", "range_txt", None, 20),
        ("Methylation FDR (gene-level)", "FDR_A", "0.00E+00", 14)]

TITLE = (f"Supplementary Table S2. The {len(d)} inverse-concordant genes (bulk RNA × methylation "
         "discovery pool) with per-gene statistics, ordered by tier. Methylation-compartment columns "
         "indicate whether each gene's inverse-concordant methylation call was supported by mCSEA "
         "promoter-region enrichment (promoter-anchored, consistent with the promoter-silencing model) "
         "or only by the gene-level Stouffer composite over promoter and gene-body probes "
         "(compartment-ambiguous, since gene-body methylation often correlates positively with "
         "transcription). “Survives promoter-only filter” reports the sensitivity analysis in "
         "which the filter is restricted to mCSEA-significant promoter regions. The reported gene-level "
         "methylation log-fold-change is the unweighted arithmetic mean of the probe log-fold-changes; "
         "the 1/SE-weighted mean (the estimator whose Wald statistic is the equal-weight Stouffer "
         "combination used for the p-value), the inverse-variance weighted mean and the per-gene probe "
         "range are given alongside so the sensitivity of each effect size to the weighting choice is "
         "visible. Weighting columns come from the 05_combined pan-tissue stratum; genes with no probe "
         "there are left blank.")

wb = load_workbook(SRC)
del wb["Supplementary_Table_S2"]
ws = wb.create_sheet("Supplementary_Table_S2")

HDR = PatternFill("solid", start_color="1F4E78")
HF = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY = Font(name="Arial", size=10)
T1F = PatternFill("solid", start_color="FFF2CC")
AUXF = PatternFill("solid", start_color="FCE4D6")

ws.cell(row=1, column=1, value=TITLE).font = Font(name="Arial", bold=True, size=10)
ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
ws.row_dimensions[1].height = 108

for j, (title, _, _, width) in enumerate(COLS, 1):
    c = ws.cell(row=3, column=j, value=title)
    c.fill, c.font = HDR, HF
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.column_dimensions[get_column_letter(j)].width = width
ws.row_dimensions[3].height = 46

for i, (_, row) in enumerate(d.iterrows(), start=4):
    fill = T1F if row.gene in TIER1 else (AUXF if row.gene in AUX else None)
    for j, (_, key, fmt, _) in enumerate(COLS, 1):
        v = row[key]
        if isinstance(v, (np.bool_, bool)):
            v = "Yes" if v else "No"
        elif pd.isna(v):
            v = None
        elif isinstance(v, (np.integer,)):
            v = int(v)
        elif isinstance(v, (np.floating,)):
            v = float(v)
        c = ws.cell(row=i, column=j, value=v)
        c.font = BODY
        if fmt and v is not None:
            c.number_format = fmt
        if fill:
            c.fill = fill
ws.freeze_panes = "B4"

wb.save(DST)
print(f"\nwrote {DST}")
print(f"  Supplementary_Table_S2: {len(d)} genes x {len(COLS)} columns")
