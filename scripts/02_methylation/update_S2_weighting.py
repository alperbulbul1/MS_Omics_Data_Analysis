#!/usr/bin/env python3
"""Add the gene-level methylation weighting columns promised in Methods and in the reply to
Reviewer 1 point 8, to Supplementary Table S2.

Columns added, all from the pan-tissue combined stratum (05_combined), which is the stratum whose
methylation statistics the manuscript reports:

Reviewer 1's point 8 offered two acceptable remedies: weight the effect size the same way as the
significance calculation, or report it as a range. We take the second, so the unweighted arithmetic
mean remains the reported effect size and the alternatives are given here for comparison.

  CpG probes (n)                       number of 450K probes aggregated for the gene
  Methylation logFC (unweighted mean)  the reported effect size, as in the original submission
  Methylation logFC (1/SE-weighted)    the estimator whose Wald statistic is the equal-weight
                                       Stouffer combination used for the p-value
  Methylation logFC (inverse-variance) belongs to a different (w = 1/SE) test; comparison only
  Probe logFC range (min / max)        the range the reviewer suggested reporting
  Methylation FDR (gene-level)         BH over the unweighted Stouffer p-values

Genes of the 82-gene pool that carry no probe in the combined stratum are left blank.
"""
import pandas as pd
import numpy as np
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = "__MS_GEO_ROOT__/Poster_v2/Supplementary_Tables_IJMS_v2.xlsx"
W = "__MS_GEO_ROOT__/Methylation/results/15_genelevel_weighting_corrected.tsv"

w = pd.read_csv(W, sep="\t")
w = w[w.stratum == "05_combined"].set_index("gene")
w["rho_mean"] = np.where(w.n_probes > 1,
                         w.rho_sum / (w.n_probes * (w.n_probes - 1) / 2), np.nan)
w["range_txt"] = [f"{lo:+.3f} … {hi:+.3f}" if n > 1 else "single probe"
                  for lo, hi, n in zip(w.min_logFC, w.max_logFC, w.n_probes)]

COLS = [("CpG probes (n)", "n_probes", "0"),
        ("Methylation logFC (unweighted mean, reported)", "mean_logFC", "0.0000"),
        ("Methylation logFC (1/SE-weighted)", "se_logFC", "0.0000"),
        ("Methylation logFC (inverse-variance)", "ivw_logFC", "0.0000"),
        ("Probe logFC range", "range_txt", None),
        ("Methylation FDR (gene-level)", "FDR_A", "0.00E+00")]

wb = load_workbook(SRC)
ws = wb["Supplementary_Table_S2"]
hdr = 3
first = ws.max_column + 1

HDR = PatternFill("solid", start_color="1F4E78")
HF = Font(name="Arial", bold=True, color="FFFFFF", size=10)
LOST = PatternFill("solid", start_color="FCE4E4")
KEPT = PatternFill("solid", start_color="C6EFCE")

n_found = n_lost = 0
for k, (title, col, fmt) in enumerate(COLS):
    j = first + k
    c = ws.cell(row=hdr, column=j, value=title)
    c.fill = HDR; c.font = HF
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = 22
    for i in range(hdr + 1, ws.max_row + 1):
        g = ws.cell(row=i, column=1).value
        if g is None or g not in w.index:
            continue
        v = w.loc[g, col]
        cell = ws.cell(row=i, column=j, value=(None if pd.isna(v) else
                                               (int(v) if col == "n_probes" else v)))
        cell.font = Font(name="Arial", size=10)
        if fmt:
            cell.number_format = fmt

for i in range(hdr + 1, ws.max_row + 1):
    g = ws.cell(row=i, column=1).value
    if g is None:
        continue
    if g in w.index:
        n_found += 1

title = str(ws.cell(row=1, column=1).value).rstrip()
ws.cell(row=1, column=1, value=title +
        " The reported gene-level methylation log-fold-change is the unweighted arithmetic mean of"
        " the probe log-fold-changes. The 1/SE-weighted mean (the estimator whose Wald statistic is"
        " the equal-weight Stouffer combination used for the p-value), the inverse-variance weighted"
        " mean and the per-gene probe range are given alongside, so that the sensitivity of each"
        " effect size to the weighting choice is visible.")
wb.save(SRC)
print(f"wrote {SRC}")
print(f"  {len(COLS)} columns added to Supplementary_Table_S2")
print(f"  {n_found} of the 82 pool genes have combined-stratum probes")

