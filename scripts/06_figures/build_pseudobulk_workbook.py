#!/usr/bin/env python3
"""Assemble all donor-level pseudobulk results into one reviewer-ready Excel workbook."""
import pandas as pd, numpy as np, os
from statsmodels.stats.multitest import multipletests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

D="__MS_GEO_ROOT__/Poster_v2/figures/pseudobulk_proper"
F="__MS_GEO_ROOT__/Poster_v2/figures"
OUT="__MS_GEO_ROOT__/Poster_v2/Pseudobulk_DonorLevel_Results.xlsx"
BH=lambda p: multipletests(p,method='fdr_bh')[1]

T1=['ITGB2','LXN','CD79B','IKZF1','SH3BP4']; SUG=['HLA-E']
AUX=['CASP6','CASP8','DGKQ','MX1','IFIT1','NUP210','RUNX3']
EXTRA=['CTSZ','CHL1','THRB','ITGAL','IFI44L','RPAP2','SLAMF1','PCNP','STAT3','TYK2','ICAM1','MOSPD3','FOXP3']
PANEL=T1+SUG+AUX+EXTRA
def tier(g): return 'Tier-1' if g in T1 else ('suggestive' if g in SUG else ('Tier-2 aux' if g in AUX else 'other panel'))

# ---------------- load ----------------
print("loading S1 (summed raw counts)...",flush=True)
s1=pd.read_csv(f"{D}/pseudobulk_muscat_style.csv")
bio=pd.read_csv(f"{D}/gene_biotypes.csv",index_col=0).biotype
s1['biotype']=s1.gene.map(bio).fillna('unknown')
print("loading S2/S3 (normalised aggregation)...",flush=True)
s23=pd.read_csv(f"{D}/pseudobulk_norm_compare.csv")
s23['biotype']=s23.gene.map(bio).fillna('unknown')

# unify
a=s1.rename(columns={'analysis':'design'})[['design','engine','cell_type','gene','logFC','PValue','FDR_local','FDR_global','biotype','n_ms','n_hc','n_pairs']]
a['aggregation']='S1_sum_raw_counts'
b=s23.rename(columns={'scheme':'design'})[['design','cell_type','gene','logFC','PValue','FDR_local','FDR_global','biotype','n_ms','n_hc','n_pairs']]
b['engine']='limma_trend'
b['aggregation']=np.where(b.design.str.startswith('S2'),'S2_mean_CP10K','S3_sum_CP10K_INVALID')
allres=pd.concat([a,b],ignore_index=True)
allres['key']=allres.aggregation+" | "+allres.design+" | "+allres.engine

# panel-scope FDR within each design x engine
allres['FDR_panel']=np.nan
for k,idx in allres.groupby('key').groups.items():
    sub=allres.loc[idx]
    p=sub[sub.gene.isin(PANEL)]
    if len(p): allres.loc[p.index,'FDR_panel']=BH(p.PValue.values)

# ---------------- sheets ----------------
sheets={}

# 1) design summary
rows=[]
for k,s in allres.groupby('key'):
    p=s[s.gene.isin(PANEL)]
    rows.append(dict(Aggregation=s.aggregation.iloc[0], Design=s.design.iloc[0], Engine=s.engine.iloc[0],
        Cell_types=s.cell_type.nunique(), Tests=len(s),
        Genes_FDRglobal_lt005=int((s.FDR_global<0.05).sum()),
        Genes_FDRlocal_lt005=int((s.FDR_local<0.05).sum()),
        Panel_tests=len(p), Panel_FDR_lt005=int((p.FDR_panel<0.05).sum()),
        MS_donors=int(s.n_ms.max()), HC_donors=int(s.n_hc.max())))
sheets['01_Design_summary']=pd.DataFrame(rows).sort_values(['Aggregation','Design','Engine'])

# 2) tier-1 headline: best result per gene per aggregation
rows=[]
for g in T1+SUG+['MOSPD3','IFI44L','MX1','DGKQ']:
    for agg,s in allres[allres.gene==g].groupby('aggregation'):
        if not len(s): continue
        # select the row that determines significance: smallest panel-scope FDR (tie-break on p)
        s=s.sort_values(['FDR_panel','PValue'], na_position='last')
        r=s.iloc[0]
        invalid = agg.endswith('INVALID')
        rows.append(dict(Tier=tier(g), Gene=g, Aggregation=agg,
            Scheme_valid="NO - do not cite" if invalid else "yes",
            Best_cell_type=r.cell_type,
            Design=r.design, Engine=r.engine, logFC=r.logFC, P_value=r.PValue,
            FDR_local=r.FDR_local, FDR_global=r.FDR_global, FDR_panel=r.FDR_panel,
            MS_donors=r.n_ms, HC_donors=r.n_hc,
            Significant_panel=("n/a (invalid scheme)" if invalid
                               else ("YES" if r.FDR_panel<0.05 else "no"))))
sheets['02_Candidates_best']=pd.DataFrame(rows).sort_values(['Scheme_valid','Tier','Gene'],
                                                            ascending=[True,True,True])

# 3) aggregation comparison (panel FDR side by side)
piv=allres[allres.gene.isin(PANEL)].groupby(['gene','aggregation']).FDR_panel.min().unstack()
pv =allres[allres.gene.isin(PANEL)].groupby(['gene','aggregation']).PValue.min().unstack()
cmp=pd.DataFrame({'Gene':piv.index,'Tier':[tier(g) for g in piv.index]})
for c in piv.columns:
    cmp[f'minP_{c}']=pv[c].values; cmp[f'panelFDR_{c}']=piv[c].values
sheets['03_Aggregation_comparison']=cmp.sort_values(['Tier','Gene'])

# 4) full candidate results (valid schemes only)
full=allres[allres.gene.isin(PANEL) & (allres.aggregation!='S3_sum_CP10K_INVALID')].copy()
full['Tier']=full.gene.map(tier)
sheets['04_Candidates_full_valid']=full[['Tier','gene','cell_type','aggregation','design','engine',
    'logFC','PValue','FDR_local','FDR_global','FDR_panel','n_ms','n_hc','n_pairs']]\
    .sort_values(['Tier','gene','PValue']).rename(columns={'gene':'Gene','cell_type':'Cell_type'})

# 5) S3 artifact demonstration
s3=allres[(allres.aggregation=='S3_sum_CP10K_INVALID')&(allres.gene.isin(PANEL))]
art=s3[s3.design=='S3_sumCP10K_coarse8'].sort_values('PValue').head(40)
sheets['05_S3_artifact_demo']=art[['gene','cell_type','logFC','PValue','FDR_local','FDR_global','FDR_panel']]\
    .rename(columns={'gene':'Gene','cell_type':'Cell_type'})

# 6) differential abundance
das=[]
for tag in ['coarse','fine']:
    p=f"{D}/DA_{tag}.csv"
    if os.path.exists(p):
        d=pd.read_csv(p); d.insert(0,'Resolution',tag); das.append(d)
if das:
    da=pd.concat(das,ignore_index=True)
    keep=[c for c in ['Resolution','cell_type','logFC','AveExpr','t','P.Value','adj.P.Val'] if c in da.columns]
    sheets['06_Differential_abundance']=da[keep]

# 7) variance components
M=pd.read_csv(f"{D}/PB_lognorm_matrix.csv",index_col=0)
C=pd.read_csv(f"{D}/PB_lognorm_coldata.csv"); C=C[C['sample'].isin(M.columns)]
rows=[]
for g in T1+SUG:
    if g not in M.index: continue
    for ct in ['t_cells','monocytes','b_cells','nk_cells']:
        s=C[C.cell_type==ct]
        if not len(s): continue
        v=M.loc[g,s['sample'].values].astype(float).values; nc=s.n_cells.values
        dsd=float(np.std(v,ddof=1)); cse=float(np.mean(np.sqrt(max(v.mean(),1e-9)/nc)))
        rows.append(dict(Gene=g,Cell_type=ct,Donors=len(v),Mean_cells_per_donor=int(nc.mean()),
            Between_donor_SD=dsd,Cell_sampling_SE=cse,Ratio_SE_over_SD=cse/max(dsd,1e-9)))
sheets['07_Variance_components']=pd.DataFrame(rows)

# 8) cell-level vs donor-level
p=f"{F}/scrna_PSEUDOBULK_COMPARISON.tsv"
if os.path.exists(p):
    cd=pd.read_csv(p,sep='\t')
    cd=cd[cd.gene.isin(PANEL)].copy(); cd.insert(0,'Tier',cd.gene.map(tier))
    sheets['08_CellLevel_vs_Donor']=cd.sort_values(['Tier','cell_wilcox_FDR_OLD'])

# 9) top genome-wide hits (pipeline sensitivity evidence)
tops=[]
for k,s in allres[allres.aggregation=='S1_sum_raw_counts'].groupby('key'):
    t=s[s.FDR_global<0.05].nsmallest(15,'FDR_global')
    if len(t): tops.append(t.assign(Source=k))
if tops:
    tp=pd.concat(tops,ignore_index=True)
    sheets['09_TopGenomewide_S1']=tp[['Source','cell_type','gene','biotype','logFC','PValue','FDR_global']]\
        .rename(columns={'gene':'Gene','cell_type':'Cell_type'})

# 10) biotype summary
sheets['10_Gene_biotypes']=bio.value_counts().rename_axis('Biotype').reset_index(name='N_genes')

# ---------------- write ----------------
wb=Workbook(); wb.remove(wb.active)
HDR=PatternFill('solid',start_color='1F4E78'); HF=Font(name='Arial',bold=True,color='FFFFFF',size=10)
BF=Font(name='Arial',size=10); TF=Font(name='Arial',bold=True,size=12)
SIG=PatternFill('solid',start_color='C6EFCE'); WARN=PatternFill('solid',start_color='FFC7CE')
thin=Side(style='thin',color='BFBFBF'); BRD=Border(bottom=thin)

# README
ws=wb.create_sheet('00_README')
readme=[
 ("Donor-level pseudobulk differential-state analysis - Kaufmann PBMC (GSE144744)",'title'),
 ("",''),
 ("Cohort: 31 MS / 31 HC donors, 497,705 cells, matched MS-HC pairs multiplexed per 10x run (31 pairs).",''),
 ("Design: paired ~batch_pair + condition. Pairs are sex-matched 31/31; mean age 38.48 (HC) vs 38.52 (MS),",''),
 ("so the pair term absorbs sex and age - no extra covariates needed.",''),
 ("",''),
 ("Raw counts were recovered exactly from the deposited log-normalised matrix:",''),
 ("   count = expm1(lognorm) * nCount_RNA / 1e4   -- verified 100% integer; per-cell expm1 sums = 10,000;",''),
 ("   reconstructed per-cell totals equal the recorded nCount_RNA exactly.",''),
 ("",''),
 ("AGGREGATION SCHEMES",'h'),
 ("S1_sum_raw_counts     SUM of raw integer UMIs per donor x cell type -> edgeR-QL and limma-voom (muscat standard)",''),
 ("S2_mean_CP10K         MEAN of per-cell CP10K, log2 -> limma-trend (equal weight per cell)",''),
 ("S3_sum_CP10K_INVALID  SUM of per-cell CP10K - REJECTED, see sheet 05: cell number re-enters as an",''),
 ("                      uncontrolled covariate (21% of genes reach FDR<0.05; effects track cell proportion)",''),
 ("",''),
 ("RESOLUTIONS: 8 coarse lineages / 25 published subsets / whole-PBMC (one sample per donor)",''),
 ("FILTERS: min 10 cells per donor x cell type; filterByExpr or >=50% non-zero; protein-coding only (mygene.info)",''),
 ("FDR SCOPES reported separately: local (within cell type), global (all gene x cell type), panel (26 candidates)",''),
 ("",''),
 ("KEY RESULTS",'h'),
 ("IKZF1   panel FDR 0.047 (whole-PBMC, S2 mean-CP10K), logFC -0.078; direction concordant with the bulk RNA",''),
 ("        down-call. Not significant under S1 (0.405). Most defensible candidate signal.",''),
 ("HLA-E   panel FDR 0.047 (cDC, S1 voom), logFC +0.212 - direction DISCORDANT with the bulk down-call.",''),
 ("MOSPD3  panel FDR 0.043 (S1) / 0.047 (S2), whole-PBMC - but not one of the 17 named candidates.",''),
 ("ITGB2   null under every scheme (panel FDR 0.42-0.44). Its tier-1 anchoring rests on bulk RNA (q=3.4e-4),",''),
 ("        promoter methylation (mCSEA q=0.016) and UK Biobank plasma proteomics (q=1.5e-17), not single-cell.",''),
 ("No candidate is robust across both valid aggregation schemes.",''),
 ("",''),
 ("WHY MORE POWER IS NOT AVAILABLE (sheet 07)",'h'),
 ("Cell-sampling SE is only 8-34% of between-donor SD, so the contrast is donor-limited: effective n = 62",''),
 ("donors, not 497,705 cells. Mixed models / kNN / cell-selection act on the cell dimension and cannot help.",''),
 ("Cell-type composition does not differ (sheet 06, all FDR > 0.19), so composition shift is not an explanation.",''),
 ("",''),
 ("SHEETS",'h'),
 ("01 Design summary | 02 Candidate best results | 03 Aggregation comparison | 04 Full candidate results",''),
 ("05 S3 artifact demo | 06 Differential abundance | 07 Variance components | 08 Cell- vs donor-level",''),
 ("09 Top genome-wide hits | 10 Gene biotypes",''),
]
for i,(txt,kind) in enumerate(readme,1):
    c=ws.cell(row=i,column=1,value=txt)
    c.font=TF if kind=='title' else (Font(name='Arial',bold=True,size=10) if kind=='h' else BF)
ws.column_dimensions['A'].width=120

for name,df in sheets.items():
    ws=wb.create_sheet(name[:31])
    for r in dataframe_to_rows(df,index=False,header=True): ws.append(r)
    for c in ws[1]: c.fill=HDR; c.font=HF; c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.freeze_panes='A2'
    hdr=[c.value for c in ws[1]]
    for j,h in enumerate(hdr,1):
        L=get_column_letter(j)
        ws.column_dimensions[L].width=min(max(12,len(str(h))+3),30)
        for i in range(2,ws.max_row+1):
            cell=ws.cell(row=i,column=j); cell.font=BF; cell.border=BRD
            v=cell.value
            if isinstance(v,float):
                if h and ('P_value' in str(h) or 'PValue' in str(h) or str(h).startswith('minP') or 'P.Value' in str(h)):
                    cell.number_format='0.00E+00'
                elif h and ('FDR' in str(h) or 'adj.P' in str(h)):
                    cell.number_format='0.0000'
                    if v<0.05: cell.fill=SIG
                elif h and ('logFC' in str(h) or 'SD' in str(h) or 'SE' in str(h) or 'Ratio' in str(h)):
                    cell.number_format='0.0000'
                else: cell.number_format='0.000'
            if h=='Significant_panel' and v=='YES': cell.fill=SIG
            if h=='Scheme_valid' and isinstance(v,str) and v.startswith('NO'): cell.fill=WARN
    # grey out / flag rows from the invalid S3 scheme so they cannot be misread as findings
    if 'Aggregation' in hdr:
        ja=hdr.index('Aggregation')+1
        for i in range(2,ws.max_row+1):
            if str(ws.cell(row=i,column=ja).value).endswith('INVALID'):
                for j in range(1,len(hdr)+1):
                    ws.cell(row=i,column=j).fill=WARN
                    ws.cell(row=i,column=j).font=Font(name='Arial',size=10,italic=True,color='9C0006')
    if name=='05_S3_artifact_demo':
        for i in range(2,ws.max_row+1):
            for j in range(1,len(hdr)+1): ws.cell(row=i,column=j).fill=WARN
wb.save(OUT)
print(f"\nwrote {OUT}")
for n,d in sheets.items(): print(f"  {n:32s} {len(d):6d} rows x {len(d.columns)} cols")
