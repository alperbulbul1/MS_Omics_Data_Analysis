import gzip
import os
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension


ROOT = Path("__MS_GEO_ROOT__")
METH_METADATA = ROOT / "Methylation_Data" / "Combined_Methylation_Metadata.csv"
EXPR_METADATA = ROOT / "Expression_Data" / "Combined_Metadata.csv"
OUTPUT = ROOT / "Used_Methylation_and_RNAseq_Dataset_Inventory.xlsx"


def trim(text: str, limit: int = 2500) -> str:
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def base_gse(dataset: str) -> str:
    return str(dataset).split("__", 1)[0]


def normalize_for_match(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "").lower()).strip()


def classify_case_control(text: str) -> str:
    text = f" {normalize_for_match(text)} "
    control_patterns = [
        r"\bhealthy\b",
        r"\bhealthy control\b",
        r"\bhealthy donor\b",
        r"\bhealthy subject\b",
        r"\bnormal control\b",
        r"\bnon-inflammatory\b",
        r"\bnon[- ]ms\b",
        r"disease status:\s*(healthy|control|h)\b",
        r"disease state:\s*(healthy|control|h)\b",
        r"disease:\s*(healthy|control|hc|h)\b",
        r"diagnosis:\s*(healthy|control|h)\b",
        r"diagnosis_ms:\s*(healthy|control|h)\b",
        r"(^|[^a-z0-9])hc\d",
        r"(^|[^a-z0-9])control([^-a-z0-9]|$)",
    ]
    case_patterns = [
        r"\bmultiple sclerosis\b",
        r"\bclinically stable ms\b",
        r"\bms patient(s)?\b",
        r"\brrms\b",
        r"\bspms\b",
        r"\bppms\b",
        r"\bcis\b",
        r"\brelapsing[- ]remitting\b",
        r"\bsecondary progressive\b",
        r"\bprimary progressive\b",
        r"disease status:\s*case\b",
        r"disease status:\s*(ms|multiple sclerosis|ms case)\b",
        r"disease state:\s*case\b",
        r"disease state:\s*(ms|multiple sclerosis)\b",
        r"disease:\s*ms\b",
        r"diagnosis:\s*ms\b",
        r"diagnosis_ms:\s*ms\b",
        r"(^|[^a-z0-9])ms\d",
        r"(^|[^a-z0-9])sms\d",
    ]

    is_control = any(re.search(pattern, text) for pattern in control_patterns)
    is_case = any(re.search(pattern, text) for pattern in case_patterns)

    if is_case and not is_control:
        return "MS"
    if is_control and not is_case:
        return "HC"
    return "Other"


def extract_other_group_labels(text: str) -> list[str]:
    text = f" {text.lower()} "
    patterns = [
        ("RRMS", [" rrms ", " relapsing-remitting "]),
        ("SPMS", [" spms ", " secondary progressive "]),
        ("PPMS", [" ppms ", " primary progressive "]),
        ("CIS", [" cis "]),
        ("Relapse", [" relapse "]),
        ("Remission", [" remission "]),
        ("Treated", [" treatment: ", " treated "]),
        ("Untreated", [" untreated ", " wash_out "]),
        ("DMF", [" dimethyl fumarate ", " dmf "]),
        ("IFNb", [" ifnb ", " interferon", " avonex ", " rebif "]),
        ("Responder", [" responder "]),
        ("Non-responder", [" non-responder ", " nonresponder "]),
        ("Smoker", [" smoker ", " smoking_status: s", " ever smoker "]),
        ("Non-smoker", [" nonsmoker ", " smoking_status: ns", " never smoker "]),
        ("Whole blood", [" whole blood ", " peripheral blood ", " pbl "]),
        ("PBMC", [" pbmc ", " mononuclear "]),
        ("BAL", [" bronchoalveolar lavage ", " bal_"]),
        ("CD4", [" cd4 "]),
        ("CD8", [" cd8 "]),
        ("CD14", [" cd14 "]),
        ("CD19", [" cd19 "]),
    ]
    labels = [label for label, terms in patterns if any(term in text for term in terms)]
    return labels or ["Other"]


def extract_cell_tissue_labels(text: str) -> list[str]:
    text = f" {text.lower()} "
    patterns = [
        ("Whole blood", [" whole blood ", " peripheral blood ", " pbl "]),
        ("PBMC", [" pbmc ", " mononuclear "]),
        ("BAL", [" bronchoalveolar lavage ", " bal_"]),
        ("Brain / WM", [" white matter ", " brain ", " frontal lobe ", " lesion ", " cortex "]),
        ("CD4 T cells", [" cd4+", " cd4 t", " cd4 t cells", " cd4 t cell"]),
        ("CD8 T cells", [" cd8+", " cd8 t", " cd8 t cells", " cd8 t cell"]),
        ("CD14 monocytes", [" cd14", " monocyte "]),
        ("CD19 B cells", [" cd19", " b cells ", " b cell "]),
        ("B cells", [" b cells ", " b cell "]),
        ("T cells", [" t cells ", " t cell "]),
        ("Dendritic cells", [" dendritic ", " toldc ", " mdc "]),
        ("CSF", [" cerebrospinal fluid ", " csf "]),
    ]
    labels = [label for label, terms in patterns if any(term in text for term in terms)]
    return labels or ["Unspecified"]


def extract_treatment_labels(text: str) -> list[str]:
    text = f" {text.lower()} "
    patterns = [
        ("Untreated", [" untreated ", " treatment: untreated ", " wash_out "]),
        ("IFNb", [" ifnb ", " interferon", " avonex ", " rebif "]),
        ("DMF", [" dimethyl fumarate ", " dmf "]),
        ("Methylprednisolone", [" methylprednisolone "]),
        ("Gilenya / Fingolimod", [" gilenya ", " fingolimod "]),
        ("Other treatment", [" treatment: other "]),
        ("Responder", [" responder "]),
        ("Non-responder", [" non-responder ", " nonresponder "]),
        ("Smoker", [" smoking_status: s", " smoker ", " ever smoker "]),
        ("Non-smoker", [" smoking_status: ns", " nonsmoker ", " never smoker "]),
        ("Relapse", [" relapse "]),
        ("Remission", [" remission "]),
    ]
    labels = [label for label, terms in patterns if any(term in text for term in terms)]
    return labels or ["Not reported"]


def extract_case_control_indicators(text: str) -> list[str]:
    lowered = text.lower()
    indicators = []
    patterns = [
        r"disease:\s*[^;]+",
        r"disease state:\s*[^;]+",
        r"disease status:\s*[^;]+",
        r"diagnosis:\s*[^;]+",
        r"diagnosis_ms:\s*[^;]+",
        r"group:\s*[^;]+",
        r"relapse_remission:\s*[^;]+",
        r"treatment:\s*[^;]+",
        r"treatment_orig:\s*[^;]+",
        r"smoking_status:\s*[^;]+",
    ]
    normalized = lowered.replace("  ", " ")
    for pattern in patterns:
        for match in re.findall(pattern, normalized):
            indicators.append(trim(match, 80))
    return indicators or ["title/characteristics text"]


def case_control_rule_text() -> str:
    return (
        "Case = samples annotated with MS-like labels in GEO metadata "
        "(e.g. multiple sclerosis, MS case, diagnosis: MS, RRMS, SPMS, PPMS). "
        "Control = samples annotated as healthy/control/HC/non-inflammatory. "
        "Samples lacking a clean case/control label or belonging to other strata "
        "(e.g. relapse, treatment-only, responder groups) were counted as Other and excluded "
        "from the pairwise MS vs HC analysis when not directly comparable."
    )


def assay_class_from_types(series_types: list[str]) -> str:
    joined = " | ".join(series_types).lower()
    if "methylation" in joined:
        return "Methylation"
    if "expression profiling by high throughput sequencing" in joined or "rna-seq" in joined:
        return "RNA-seq"
    if "expression profiling by array" in joined:
        return "Microarray"
    return "Other"


def detect_cell_tissue_key(text: str) -> str:
    text = f" {normalize_for_match(text)} "
    mapping = [
        ("whole_blood", [" whole blood ", " wholeblood ", " peripheral blood ", " tissue: blood "]),
        ("pbmc", [" pbmc ", " mononuclear ", " peripheral blood mononuclear cells "]),
        ("bronchoalveolar_lavage", [" bronchoalveolar lavage ", " bal cells ", " bronchoalveolar "]),
        ("brain_wm", [" white matter ", " frontal lobe ", " brain ", " cortex ", " corpus callosum "]),
        ("cd4_tcell", [" cd4t ", " cd4+ ", " cd4 t", " cd4 t cell", " cd4 t cells"]),
        ("cd8_tcell", [" cd8t ", " cd8+ ", " cd8 t", " cd8 t cell", " cd8 t cells"]),
        ("cd14_monocyte", [" cd14 ", " monocyte "]),
        ("cd19_bcell", [" cd19 ", " b cell ", " b cells "]),
        ("dendritic_cell", [" dendritic ", " toldc ", " mdc "]),
        ("csf", [" cerebrospinal fluid ", " csf "]),
    ]
    for label, terms in mapping:
        if any(term in text for term in terms):
            return label
    return "unspecified"


def resolve_soft_path(gse_id: str, omics: str) -> Path | None:
    if omics == "Methylation":
        p = ROOT / "Methylation_Data" / gse_id / f"{gse_id}_family.soft.gz"
        return p if p.exists() else None

    folder = ROOT / "Expression_Data" / gse_id / f"{gse_id}_family.soft.gz"
    root = ROOT / "Expression_Data" / f"{gse_id}_family.soft.gz"
    if folder.exists():
        return folder
    if root.exists():
        return root
    return None


def parse_soft(soft_path: Path) -> dict:
    series = {
        "title": "",
        "summary": [],
        "overall_design": "",
        "pubmed_ids": [],
        "series_types": [],
        "series_platform_ids": [],
        "platforms": {},
    }
    samples = {}
    current_sample = None
    current_platform = None
    in_table = False
    prefixes = (
        "!Sample_title = ",
        "!Sample_characteristics_ch1 = ",
        "!Sample_source_name_ch1 = ",
        "!Sample_description = ",
    )

    with gzip.open(soft_path, "rt", encoding="utf-8", errors="ignore") as handle:
        for raw_line in handle:
            line = raw_line.rstrip("\n")
            stripped = line.strip()

            if stripped.lower() == "!sample_table_begin":
                in_table = True
                continue
            if stripped.lower() == "!sample_table_end":
                in_table = False
                continue
            if in_table:
                continue

            if stripped.startswith("^SAMPLE = "):
                current_sample = stripped.split("=", 1)[1].strip()
                samples[current_sample] = []
                continue

            if stripped.startswith("^PLATFORM = "):
                current_platform = stripped.split("=", 1)[1].strip()
                series["platforms"].setdefault(
                    current_platform,
                    {"title": "", "technology": "", "manufacturer": ""},
                )
                continue

            if current_sample is not None:
                if stripped.startswith("^"):
                    current_sample = None
                else:
                    for prefix in prefixes:
                        if stripped.startswith(prefix):
                            samples[current_sample].append(stripped.split("=", 1)[1].strip())
                            break

            if current_platform is not None:
                if stripped.startswith("^") and not stripped.startswith("^PLATFORM = "):
                    current_platform = None
                elif stripped.startswith("!Platform_title = "):
                    series["platforms"][current_platform]["title"] = stripped.split("=", 1)[1].strip()
                elif stripped.startswith("!Platform_technology = "):
                    series["platforms"][current_platform]["technology"] = stripped.split("=", 1)[1].strip()
                elif stripped.startswith("!Platform_manufacturer = "):
                    series["platforms"][current_platform]["manufacturer"] = stripped.split("=", 1)[1].strip()

            if stripped.startswith("!Series_title = "):
                series["title"] = stripped.split("=", 1)[1].strip()
            elif stripped.startswith("!Series_summary = "):
                series["summary"].append(stripped.split("=", 1)[1].strip())
            elif stripped.startswith("!Series_overall_design = "):
                series["overall_design"] = stripped.split("=", 1)[1].strip()
            elif stripped.startswith("!Series_pubmed_id = "):
                series["pubmed_ids"].append(stripped.split("=", 1)[1].strip())
            elif stripped.startswith("!Series_type = "):
                series["series_types"].append(stripped.split("=", 1)[1].strip())
            elif stripped.startswith("!Series_platform_id = "):
                series["series_platform_ids"].append(stripped.split("=", 1)[1].strip())

    sample_records = []
    for sample_id, parts in samples.items():
        text = " ".join(parts)
        primary = classify_case_control(text)
        other_labels = extract_other_group_labels(text)
        sample_records.append(
            {
                "sample_id": sample_id,
                "raw_text": text,
                "primary_group": primary,
                "other_labels": other_labels,
                "case_control_indicators": extract_case_control_indicators(text),
                "cell_tissue_labels": extract_cell_tissue_labels(text),
                "treatment_labels": extract_treatment_labels(text),
                "cell_tissue_key": detect_cell_tissue_key(text),
            }
        )

    platform_ids = [pid for pid in series["series_platform_ids"] if pid]
    if not platform_ids and series["platforms"]:
        platform_ids = list(series["platforms"].keys())
    platform_titles = [
        series["platforms"].get(pid, {}).get("title", "").strip()
        for pid in platform_ids
        if series["platforms"].get(pid, {}).get("title", "").strip()
    ]
    platform_technologies = [
        series["platforms"].get(pid, {}).get("technology", "").strip()
        for pid in platform_ids
        if series["platforms"].get(pid, {}).get("technology", "").strip()
    ]
    platform_manufacturers = [
        series["platforms"].get(pid, {}).get("manufacturer", "").strip()
        for pid in platform_ids
        if series["platforms"].get(pid, {}).get("manufacturer", "").strip()
    ]

    return {
        "title": trim(series["title"], 300),
        "summary": trim(" ".join(series["summary"]), 2500),
        "overall_design": trim(series["overall_design"], 2500),
        "pubmed_ids": ", ".join(pid for pid in series["pubmed_ids"] if pid and pid != "NA"),
        "series_types": series["series_types"],
        "assay_class": assay_class_from_types(series["series_types"]),
        "platform_ids": platform_ids,
        "platform_titles": platform_titles,
        "platform_technologies": platform_technologies,
        "platform_manufacturers": platform_manufacturers,
        "samples": sample_records,
    }


def distribution_string(counter: Counter) -> str:
    if not counter:
        return ""
    return "; ".join(f"{k}: {v}" for k, v in counter.most_common())


def usage_status_text(
    matched_used: int,
    total_used: int,
    eligible_case: int,
    used_case: int,
    eligible_control: int,
    used_control: int,
    eligible_other: int,
    used_other: int,
) -> str:
    if total_used == 0:
        return "No analysis samples recorded"
    if matched_used != total_used:
        return f"Sample ID mismatch in GEO SOFT ({matched_used}/{total_used} matched)"
    if (
        eligible_case == used_case
        and eligible_control == used_control
        and eligible_other == used_other
        and (eligible_case + eligible_control + eligible_other) == total_used
    ):
        return "All file-verified eligible samples used"
    if eligible_case >= used_case and eligible_control >= used_control and eligible_other >= used_other:
        return (
            f"Partial file-verified use "
            f"(case {used_case}/{eligible_case}, control {used_control}/{eligible_control}, other {used_other}/{eligible_other})"
        )
    return "Used cohort derived from broader or nonstandard source annotations"


def build_rows(metadata_path: Path, omics: str) -> list[dict]:
    used_meta = pd.read_csv(metadata_path)
    used_meta["base_gse"] = used_meta["dataset"].map(base_gse)

    rows = []
    for gse_id in sorted(used_meta["base_gse"].dropna().astype(str).unique()):
        analysis_df = used_meta[used_meta["base_gse"] == gse_id].copy()
        soft_path = resolve_soft_path(gse_id, omics)
        if soft_path is None:
            continue

        parsed = parse_soft(soft_path)
        sample_records = parsed["samples"]

        study_primary = Counter(record["primary_group"] for record in sample_records)
        study_other = Counter()
        study_cell_tissue = Counter()
        study_treatment = Counter()
        study_indicators = Counter()
        for record in sample_records:
            if record["primary_group"] == "Other":
                study_other.update(record["other_labels"])
            study_cell_tissue.update(record["cell_tissue_labels"])
            study_treatment.update(record["treatment_labels"])
            study_indicators.update(record["case_control_indicators"])

        sample_lookup = {record["sample_id"]: record for record in sample_records}
        used_records = [sample_lookup[sid] for sid in analysis_df["sample_id"].astype(str) if sid in sample_lookup]
        used_cell_tissue = Counter()
        used_treatment = Counter()
        used_indicators = Counter()
        for record in used_records:
            used_cell_tissue.update(record["cell_tissue_labels"])
            used_treatment.update(record["treatment_labels"])
            used_indicators.update(record["case_control_indicators"])

        analysis_case = int((analysis_df["condition"] == "MS").sum()) if "condition" in analysis_df.columns else 0
        analysis_control = int((analysis_df["condition"] == "HC").sum()) if "condition" in analysis_df.columns else 0
        analysis_other = int(len(analysis_df) - analysis_case - analysis_control)
        matched_used = len(used_records)

        subset_keys = set()
        if "cell_type" in analysis_df.columns:
            subset_keys = {
                str(value).strip()
                for value in analysis_df["cell_type"].dropna().astype(str)
                if str(value).strip() and str(value).strip() != "unspecified"
            }
        relevant_records = sample_records
        relevant_subset = "All study samples"
        if len(subset_keys) == 1:
            subset_key = next(iter(subset_keys))
            filtered = [record for record in sample_records if record["cell_tissue_key"] == subset_key]
            if filtered:
                relevant_records = filtered
                relevant_subset = subset_key

        relevant_primary = Counter(record["primary_group"] for record in relevant_records)
        relevant_case = int(relevant_primary.get("MS", 0))
        relevant_control = int(relevant_primary.get("HC", 0))
        relevant_other = int(relevant_primary.get("Other", 0))

        rows.append(
            {
                "GSE ID": gse_id,
                "Analysis Dataset Label(s)": ", ".join(sorted(analysis_df["dataset"].astype(str).unique())),
                "Omics": omics,
                "Assay Class": parsed["assay_class"],
                "Platform ID(s)": ", ".join(parsed["platform_ids"]),
                "Platform Title(s)": "; ".join(parsed["platform_titles"]),
                "Platform Technology": "; ".join(parsed["platform_technologies"]),
                "Platform Manufacturer": "; ".join(parsed["platform_manufacturers"]),
                "Study Title": parsed["title"],
                "PMID": parsed["pubmed_ids"],
                "Series Types": " | ".join(parsed["series_types"]),
                "Study Total Samples": len(sample_records),
                "Study MS / Case": int(study_primary.get("MS", 0)),
                "Study HC / Control": int(study_primary.get("HC", 0)),
                "Study Other": int(study_primary.get("Other", 0)),
                "Study Other Group Breakdown": distribution_string(study_other),
                "Relevant Source Subset": relevant_subset,
                "Relevant Source Total Samples": len(relevant_records),
                "Relevant Source MS / Case": relevant_case,
                "Relevant Source HC / Control": relevant_control,
                "Relevant Source Other": relevant_other,
                "Case / Control Selection Rule": case_control_rule_text(),
                "Study Case / Control Indicators": distribution_string(study_indicators),
                "Study Cell / Tissue Breakdown": distribution_string(study_cell_tissue),
                "Study Treatment / Exposure Breakdown": distribution_string(study_treatment),
                "Used in Analysis": int(len(analysis_df)),
                "Used MS / Case": analysis_case,
                "Used HC / Control": analysis_control,
                "Used Other": analysis_other,
                "Used Samples Found in GEO SOFT": matched_used,
                "Used Sample ID Match Status": (
                    f"All matched ({matched_used}/{len(analysis_df)})"
                    if matched_used == len(analysis_df)
                    else f"Partial match ({matched_used}/{len(analysis_df)})"
                ),
                "Case / Control Usage Status": usage_status_text(
                    matched_used,
                    int(len(analysis_df)),
                    relevant_case,
                    analysis_case,
                    relevant_control,
                    analysis_control,
                    relevant_other,
                    analysis_other,
                ),
                "Used Case / Control Indicators": distribution_string(used_indicators),
                "Used Cell / Tissue Breakdown": distribution_string(used_cell_tissue),
                "Used Treatment / Exposure Breakdown": distribution_string(used_treatment),
                "Study Summary": parsed["summary"],
                "Study Overall Design": parsed["overall_design"],
            }
        )

    return rows


def style_sheet(ws, title: str) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    alt_fill = PatternFill("solid", fgColor="F7FBFF")
    thin = Side(style="thin", color="D0D7DE")

    ws["A1"] = title
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    for cell in ws[2]:
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx in range(3, ws.max_row + 1):
        fill = alt_fill if row_idx % 2 == 1 else None
        for cell in ws[row_idx]:
            if fill:
                cell.fill = fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 14,
        "B": 28,
        "C": 12,
        "D": 14,
        "E": 16,
        "F": 42,
        "G": 18,
        "H": 18,
        "I": 48,
        "J": 14,
        "K": 22,
        "L": 14,
        "M": 12,
        "N": 15,
        "O": 12,
        "P": 28,
        "Q": 20,
        "R": 14,
        "S": 12,
        "T": 15,
        "U": 12,
        "V": 48,
        "W": 38,
        "X": 30,
        "Y": 30,
        "Z": 14,
        "AA": 12,
        "AB": 15,
        "AC": 12,
        "AD": 14,
        "AE": 26,
        "AF": 42,
        "AG": 38,
        "AH": 30,
        "AI": 60,
        "AJ": 60,
    }
    for col_letter, width in widths.items():
        if col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].width = width
        else:
            ws.column_dimensions[col_letter] = ColumnDimension(ws, width=width)


def append_sheet(wb: Workbook, name: str, rows: list[dict], title: str) -> None:
    ws = wb.create_sheet(title=name)
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame([{"Message": "No data"}])

    ws.append([""] * len(df.columns))
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    style_sheet(ws, title)


def build_summary_sheet(wb: Workbook, methyl_rows: list[dict], expr_rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Summary"

    expr_df = pd.DataFrame(expr_rows)
    methyl_df = pd.DataFrame(methyl_rows)
    rnaseq_df = expr_df[expr_df["Assay Class"] == "RNA-seq"].copy() if not expr_df.empty else pd.DataFrame()

    summary_rows = [
        ["Used methylation datasets", len(methyl_df)],
        ["Used methylation samples", int(methyl_df["Used in Analysis"].sum()) if not methyl_df.empty else 0],
        ["Used RNA / transcriptome datasets", len(expr_df)],
        ["Used RNA / transcriptome samples", int(expr_df["Used in Analysis"].sum()) if not expr_df.empty else 0],
        ["Used RNA-seq datasets", len(rnaseq_df)],
        ["Used RNA-seq samples", int(rnaseq_df["Used in Analysis"].sum()) if not rnaseq_df.empty else 0],
        ["Output file", str(OUTPUT.name)],
    ]

    ws.append(["", ""])
    ws.append(["Metric", "Value"])
    for row in summary_rows:
        ws.append(row)

    style_sheet(ws, "Used Methylation and RNA-seq Dataset Inventory")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24


def main() -> None:
    methyl_rows = build_rows(METH_METADATA, "Methylation")
    expr_rows = build_rows(EXPR_METADATA, "Transcriptome")

    expr_df = pd.DataFrame(expr_rows)
    rnaseq_rows = expr_df[expr_df["Assay Class"] == "RNA-seq"].to_dict("records") if not expr_df.empty else []

    wb = Workbook()
    build_summary_sheet(wb, methyl_rows, expr_rows)
    append_sheet(wb, "Methylation_Used", methyl_rows, "Used Methylation Datasets")
    append_sheet(wb, "Transcriptome_Used", expr_rows, "Used Transcriptome Datasets")
    append_sheet(wb, "RNAseq_Only", rnaseq_rows, "Used RNA-seq Datasets")

    wb.save(OUTPUT)
    print(f"Saved workbook to {OUTPUT}")


if __name__ == "__main__":
    main()
